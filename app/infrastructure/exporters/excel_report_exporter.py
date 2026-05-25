"""
Exportador de reportes en formato Microsoft Excel (.xlsx).
Estructura la información en 4 hojas detalladas y estilizadas para auditorías premium.
"""

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.application.interfaces.report_exporter_interface import ReportExporterInterface
from app.application.services.recommendation_service import RecommendationService
from app.domain.entities.scan_report import ScanReport
from app.domain.value_objects.header_status import HeaderStatus
from app.shared.filename_utils import get_unique_filename

COL_RECOMENDACION = "recomendación"


class ExcelReportExporter(ReportExporterInterface):
    """
    Exportador Excel usando pandas y openpyxl con estilos corporativos.
    """

    def __init__(self, recommendation_service: RecommendationService | None = None):
        self._recs_service = recommendation_service or RecommendationService()

    def _build_resumen_df(self, reports: list[ScanReport]) -> pd.DataFrame:
        resumen_data = []
        for r in reports:
            total_eval = len(r.results)
            correctos = sum(
                1 for x in r.results if x.status in (HeaderStatus.CORRECTO, HeaderStatus.PRESENTE)
            )
            faltantes = sum(1 for x in r.results if x.status == HeaderStatus.FALTANTE)
            debiles = sum(1 for x in r.results if x.status == HeaderStatus.DEBIL)

            resumen_data.append(
                {
                    "url": r.url,
                    "final_url": r.final_url,
                    "status_code": r.status_code if r.status_code else "N/A",
                    "score": r.score,
                    "clasificación": r.classification.value,
                    "total_headers_evaluados": total_eval,
                    "headers_correctos": correctos,
                    "headers_faltantes": faltantes,
                    "headers_debiles": debiles,
                    "error": r.error_message if r.error_message else "",
                    "fecha_analisis": r.scan_date,
                }
            )
        return pd.DataFrame(resumen_data)

    def _build_detalle_df(self, reports: list[ScanReport]) -> pd.DataFrame:
        detalle_data = []
        for r in reports:
            for res in r.results:
                detalle_data.append(
                    {
                        "url": r.url,
                        "header": res.header_name,
                        "presente": "Sí" if res.present else "No",
                        "valor": res.value if res.value else "N/A",
                        "estado": res.status.value,
                        "severidad": res.severity,
                        COL_RECOMENDACION: res.recommendation,
                    }
                )
        return pd.DataFrame(detalle_data)

    def _build_recs_df(self, reports: list[ScanReport]) -> pd.DataFrame:
        recs_data = []
        for r in reports:
            recs = self._recs_service.generate_recommendations(r.url, r.results)
            for rec in recs:
                recs_data.append(
                    {
                        "url": rec["url"],
                        "prioridad": rec["prioridad"],
                        "header": rec["header"],
                        "problema": rec["problema"],
                        COL_RECOMENDACION: rec[COL_RECOMENDACION],
                    }
                )
        if recs_data:
            return pd.DataFrame(recs_data)
        return pd.DataFrame(columns=["url", "prioridad", "header", "problema", COL_RECOMENDACION])

    def _build_errores_df(self, reports: list[ScanReport]) -> pd.DataFrame:
        errores_data = []
        for r in reports:
            if r.error_message:
                errores_data.append(
                    {
                        "url": r.url,
                        "tipo_error": "Connection/SSL Error",
                        "mensaje_error": r.error_message,
                        "fecha_analisis": r.scan_date,
                    }
                )
        if errores_data:
            return pd.DataFrame(errores_data)
        return pd.DataFrame(columns=["url", "tipo_error", "mensaje_error", "fecha_analisis"])

    def _style_header_row(self, worksheet, fill, font, alignment) -> None:
        """Estiliza la primera fila (encabezados) de la hoja de cálculo."""
        worksheet.row_dimensions[1].height = 28
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = fill
            cell.font = font
            cell.alignment = alignment

    def _style_data_columns(self, worksheet, cell_font, center_alignment, left_alignment) -> None:
        """Estiliza las celdas de datos y ajusta el ancho de las columnas."""
        for col in worksheet.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = 0
            for cell in col:
                # Estilizar tipografía si no es encabezado
                if cell.row > 1:
                    cell.font = cell_font
                    if col_letter in ("C", "D", "E", "F", "G", "H", "I", "K"):
                        cell.alignment = center_alignment
                    else:
                        cell.alignment = left_alignment

                # Calcular longitud máxima para autoajuste
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)

            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    def _apply_sheet_styling(
        self,
        worksheet,
        header_fill,
        header_font,
        cell_font,
        center_alignment,
        left_alignment,
    ) -> None:
        self._style_header_row(worksheet, header_fill, header_font, center_alignment)
        self._style_data_columns(worksheet, cell_font, center_alignment, left_alignment)

    def export(self, reports: list[ScanReport], output_dir: str, base_filename: str) -> str:
        """
        Genera el reporte Excel con 4 hojas: Resumen, Detalle, Recomendaciones y Errores.
        Asegura que no se sobreescriban reportes previos.
        """
        output_path = Path(output_dir)
        output_file = get_unique_filename(output_path, base_filename, "xlsx")

        df_resumen = self._build_resumen_df(reports)
        df_detalle = self._build_detalle_df(reports)
        df_recs = self._build_recs_df(reports)
        df_errores = self._build_errores_df(reports)

        # Guardar en archivo utilizando pandas ExcelWriter con motor openpyxl
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            df_resumen.to_excel(writer, sheet_name="Resumen", index=False)
            df_detalle.to_excel(writer, sheet_name="Detalle Headers", index=False)
            df_recs.to_excel(writer, sheet_name="Recomendaciones", index=False)
            df_errores.to_excel(writer, sheet_name="Errores", index=False)

            # Paleta de colores profesionales (Azul Oscuro de Ciberseguridad)
            header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell_font = Font(name="Calibri", size=11, bold=False)
            center_alignment = Alignment(horizontal="center", vertical="center")
            left_alignment = Alignment(horizontal="left", vertical="center")

            # Aplicar estilos por hoja
            for sheet_name in ["Resumen", "Detalle Headers", "Recomendaciones", "Errores"]:
                self._apply_sheet_styling(
                    writer.sheets[sheet_name],
                    header_fill,
                    header_font,
                    cell_font,
                    center_alignment,
                    left_alignment,
                )

        return str(output_file.absolute())
